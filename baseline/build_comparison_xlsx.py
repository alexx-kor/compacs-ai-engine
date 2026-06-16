#!/usr/bin/env python3
"""Build Excel comparison table from baseline JSON evaluation artifacts."""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from typing import Any

import pandas as pd

ROOT_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT_DIR))

from core.datasets import GoldenItem  # noqa: E402
from core.evaluation_utils import (  # noqa: E402
    is_not_found_answer,
    is_unanswerable_expected,
    tokenize_overlap,
)

BASELINE_DIR = Path(__file__).resolve().parent
OUTPUT_FILE = BASELINE_DIR / "baseline_comparison.xlsx"

GOLDEN_FILE = BASELINE_DIR / "golden_set.json"
OLLAMA_PLAIN_FILE = BASELINE_DIR / "rag_answers_gpu.json"
OLLAMA_HYBRID_FILE = BASELINE_DIR / "evaluation_llama3b_gpt_judge_v3.json"
GPT_HYBRID_FILE = BASELINE_DIR / "evaluation_results_20260518_160608.json"
LLM_JUDGE_FILE = BASELINE_DIR / "llm_judge_scores.json"


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def answers_by_id(payload: Any) -> dict[int, dict[str, Any]]:
    if isinstance(payload, list):
        rows = payload
    else:
        rows = payload.get("results", [])
    return {int(row["id"]): row for row in rows}


class _MetricsCalculator:
    """Local copy of full_evaluation.MetricsCalculator (no DB bootstrap)."""

    @staticmethod
    def relevance(question: str, answer: str) -> float:
        if not question or not answer or answer.startswith("ERROR") or is_not_found_answer(answer):
            return 0.0
        question_words = tokenize_overlap(question)
        answer_words = tokenize_overlap(answer)
        if not question_words:
            return 5.0
        return min(10.0, (len(question_words & answer_words) / len(question_words)) * 12)

    @staticmethod
    def factuality(answer: str) -> float:
        if not answer or answer.startswith("ERROR"):
            return 0.0
        uncertain = ("probably", "maybe", "perhaps", "might", "could", "возможно", "вероятно", "наверное")
        sentences = re.split(r"[.!?]+", answer)
        hits = sum(1 for s in sentences if any(m in s.lower() for m in uncertain))
        return min(10.0, max(0.0, 10 - hits * 2))

    @staticmethod
    def completeness(question: str, answer: str) -> float:
        if not answer or answer.startswith("ERROR") or is_not_found_answer(answer):
            return 0.0
        question_words = tokenize_overlap(question)
        answer_words = tokenize_overlap(answer)
        if not question_words:
            return 7.0
        return min(10.0, (len(question_words & answer_words) / len(question_words)) * 10)

    @staticmethod
    def coherence(answer: str) -> float:
        if not answer or answer.startswith("ERROR"):
            return 0.0
        structure = 0.0
        if re.search(r"\d+\.", answer):
            structure += 0.4
        if re.search(r"[-*•]", answer):
            structure += 0.3
        if answer.count("\n\n") > 0:
            structure += 0.3
        connectors = (
            "поэтому",
            "следовательно",
            "во-первых",
            "например",
            "therefore",
            "thus",
            "consequently",
            "first",
            "for example",
        )
        logic = min(0.5, sum(1 for c in connectors if c in answer.lower()) * 0.1)
        return min(10.0, (structure + logic) * 10)

    @staticmethod
    def helpfulness(answer: str) -> float:
        if not answer or answer.startswith("ERROR") or is_not_found_answer(answer):
            return 0.0
        length_score = min(1.0, len(answer) / 500)
        has_instructions = any(
            w in answer.lower()
            for w in ("как", "следуйте", "выполните", "используйте", "подключ", "запуст", "how to", "follow", "use", "step")
        )
        has_example = bool(re.search(r"(example|например|sample|пример|```)", answer.lower()))
        has_structure = bool(re.search(r"^\s*\d+\.", answer, re.MULTILINE))
        score = (
            length_score * 0.25
            + (0.35 if has_instructions else 0.0)
            + (0.2 if has_example else 0.0)
            + (0.2 if has_structure else 0.0)
        )
        return score * 10

    @staticmethod
    def toxicity(answer: str) -> float:
        toxic_words = ("дурак", "идиот", "урод", "stupid", "idiot")
        return min(10.0, sum(1 for word in toxic_words if word in answer.lower()) * 2)


def compute_final_score(golden: GoldenItem, answer: str) -> float:
    """Same 0–10 formula as full_evaluation.evaluate_response."""
    calc = _MetricsCalculator
    question = golden.question
    if is_unanswerable_expected(golden.expected_answer):
        if is_not_found_answer(answer):
            return 10.0
        scores = (
            calc.relevance(question, answer) * 0.25
            + 2.0 * 0.25
            + calc.completeness(question, answer) * 0.20
            + calc.coherence(answer) * 0.15
            + calc.helpfulness(answer) * 0.15
        )
        return round(min(4.0, scores), 2)

    final_score = (
        calc.relevance(question, answer) * 0.25
        + calc.factuality(answer) * 0.25
        + calc.completeness(question, answer) * 0.20
        + calc.coherence(answer) * 0.15
        + calc.helpfulness(answer) * 0.15
    )
    if is_not_found_answer(answer):
        final_score = min(final_score, 3.0)
    if calc.toxicity(answer) > 7:
        final_score *= 0.5
    return round(final_score, 2)


def format_sources(sources: Any) -> str:
    if not sources:
        return ""
    parts: list[str] = []
    for item in sources:
        if isinstance(item, (list, tuple)) and len(item) >= 2:
            parts.append(f"{item[0]} (стр. {item[1]})")
        else:
            parts.append(str(item))
    return "; ".join(parts)


def load_llm_judge_scores() -> dict[str, dict[int, dict[str, Any]]]:
    """Load per-run LLM judge totals from llm_judge_scores.json."""
    if not LLM_JUDGE_FILE.exists():
        return {}
    payload = load_json(LLM_JUDGE_FILE)
    runs = payload.get("runs", {})
    parsed: dict[str, dict[int, dict[str, Any]]] = {}
    for run_key, by_id in runs.items():
        parsed[run_key] = {int(k): v for k, v in by_id.items() if isinstance(v, dict)}
    return parsed


def _judge_total_from_row(row: dict[str, Any]) -> float | None:
    """Extract 0–100 total from judge payload (explicit total or metric average)."""
    if "error" in row:
        return None
    if row.get("total") is not None:
        return round(float(row["total"]), 1)
    parts = [row[key] for key in ("relevance", "accuracy", "completeness", "clarity") if key in row]
    numeric = [float(v) for v in parts if isinstance(v, (int, float))]
    if numeric:
        return round(sum(numeric) / len(numeric), 1)
    return None


def judge_total(
    judge_scores: dict[str, dict[int, dict[str, Any]]],
    run_key: str,
    qid: int,
    fallback_row: dict[str, Any] | None = None,
) -> str | float:
    """Return LLM judge total (0-100) for a question."""
    row = judge_scores.get(run_key, {}).get(qid)
    if row:
        total = _judge_total_from_row(row)
        if total is not None:
            return total
    if fallback_row:
        nested = fallback_row.get("llm_judge")
        if isinstance(nested, dict):
            total = _judge_total_from_row(nested)
            if total is not None:
                return total
    return ""


def load_golden_items() -> dict[int, GoldenItem]:
    return {
        int(item["id"]): GoldenItem(
            id=int(item["id"]),
            question=str(item["question"]),
            expected_answer=str(item["expected_answer"]),
        )
        for item in load_json(GOLDEN_FILE)
    }


def build_rows() -> list[dict[str, Any]]:
    golden = load_golden_items()
    judge = load_llm_judge_scores()
    ollama_plain = answers_by_id(load_json(OLLAMA_PLAIN_FILE))
    ollama_hybrid = answers_by_id(load_json(OLLAMA_HYBRID_FILE))
    gpt_hybrid = answers_by_id(load_json(GPT_HYBRID_FILE))

    rows: list[dict[str, Any]] = []
    for qid in sorted(golden):
        gold = golden[qid]
        plain = ollama_plain.get(qid, {})
        o_hybrid = ollama_hybrid.get(qid, {})
        g_hybrid = gpt_hybrid.get(qid, {})
        plain_answer = str(plain.get("answer", ""))

        rows.append(
            {
                "№": qid,
                "Вопрос": gold.question,
                "Золотой сет": gold.expected_answer,
                "Ollama без гибридного поиска": plain_answer,
                "Оценка Ollama (без гибрида)": judge_total(judge, "ollama_plain", qid),
                "Ollama с гибридным поиском": o_hybrid.get("answer", ""),
                "Оценка Ollama (гибрид)": judge_total(judge, "ollama_hybrid", qid, o_hybrid),
                "GPT с гибридным поиском": g_hybrid.get("answer", ""),
                "Оценка GPT (гибрид)": judge_total(judge, "gpt_hybrid", qid),
                "Источники Ollama (без гибрида)": format_sources(plain.get("sources")),
                "Источники Ollama (гибрид)": format_sources(o_hybrid.get("sources")),
                "Источники GPT (гибрид)": format_sources(g_hybrid.get("sources")),
            }
        )
    return rows


def average_judge_score(run_key: str) -> float | str:
    judge = load_llm_judge_scores()
    totals = [
        t
        for row in judge.get(run_key, {}).values()
        if isinstance(row, dict) and (t := _judge_total_from_row(row)) is not None
    ]
    return round(sum(totals) / len(totals), 1) if totals else "—"


def build_metadata_sheet() -> pd.DataFrame:
    def meta_row(label: str, path: Path, extra: dict[str, Any] | None = None) -> dict[str, Any]:
        payload = load_json(path)
        cfg = payload.get("config", {}) if isinstance(payload, dict) else {}
        row: dict[str, Any] = {
            "Прогон": label,
            "Файл": path.name,
            "Модель": cfg.get("model", "llama3.2:3b (run_gpu_baseline)"),
            "LLM provider": cfg.get("llm_provider", "ollama"),
            "Чанков": cfg.get("chunks", "—"),
            "Вопросов": cfg.get("questions", 28),
            "Средний балл (локальный)": (
                payload.get("statistics", {}).get("avg_score")
                if isinstance(payload, dict)
                else "—"
            ),
            "Средний балл LLM-судья (%)": "—",
            "Гибридный поиск": extra.get("hybrid", "") if extra else "",
            "Примечание": extra.get("note", "") if extra else "",
        }
        return row

    records = [
        meta_row(
            "Золотой сет",
            GOLDEN_FILE,
            {"hybrid": "—", "note": "Эталонные ответы (reference)"},
        ),
        {
            **meta_row(
                "Ollama без гибридного поиска",
                OLLAMA_PLAIN_FILE,
                {
                    "hybrid": "нет",
                    "note": "run_gpu_baseline.py; оценка — GPT-4o-mini vs golden",
                },
            ),
            "Средний балл LLM-судья (%)": average_judge_score("ollama_plain"),
        },
        {
            **meta_row(
                "Ollama с гибридным поиском",
                OLLAMA_HYBRID_FILE,
                {
                    "hybrid": "да",
                    "note": "full_evaluation + llama3.2:3b; судья gpt-4o-mini",
                },
            ),
            "Средний балл LLM-судья (%)": average_judge_score("ollama_hybrid"),
        },
        {
            **meta_row(
                "GPT с гибридным поиском",
                GPT_HYBRID_FILE,
                {
                    "hybrid": "да",
                    "note": "full_evaluation + gpt-4o-mini; оценка судьёй",
                },
            ),
            "Средний балл LLM-судья (%)": average_judge_score("gpt_hybrid"),
        },
    ]
    return pd.DataFrame(records)


def autosize_columns(worksheet: Any, max_width: int = 80) -> None:
    from openpyxl.utils import get_column_letter

    for column_cells in worksheet.columns:
        letter = get_column_letter(column_cells[0].column)
        length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[letter].width = min(max(length + 2, 12), max_width)


def main() -> None:
    comparison_df = pd.DataFrame(build_rows())
    metadata_df = build_metadata_sheet()

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        comparison_df.to_excel(writer, sheet_name="Сравнение", index=False)
        metadata_df.to_excel(writer, sheet_name="Метаданные", index=False)

        from openpyxl.styles import Alignment

        wrapped = Alignment(wrap_text=True, vertical="top")
        ws = writer.sheets["Сравнение"]
        ws.freeze_panes = "A2"
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = wrapped
        autosize_columns(ws)

        ws_meta = writer.sheets["Метаданные"]
        ws_meta.freeze_panes = "A2"
        autosize_columns(ws_meta, max_width=50)

    print(f"Written: {OUTPUT_FILE} ({len(comparison_df)} rows)")


if __name__ == "__main__":
    main()
